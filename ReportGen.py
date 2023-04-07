import datetime, os, json, random
from tkinter import *
from tkinter import ttk
from openpyxl import Workbook
from openpyxl.styles import *

# Main window's sizes
WIN_Y = 10
WIN_X = 100
WIN_W = 300
WIN_H = 300

YEARS = ["2020", "2021", "2022", "2023", "2024", "2025", "2026", "2027", "2028", "2029", "2030" ]
MONTHS = ["января", "февраля", "марта", "апреля", "мая", "июня", "июля", "августа", "сентября", "октября",  "ноября", "декабря"]

# Number of weeks with "one a week" tasks
NUM_WEEKS = [2, 6, 10, 14, 19, 23, 27, 32, 37, 41, 46, 51]

ROOT = os.path.dirname(__file__)
DATA_FILES = (os.path.join(ROOT, "data_files"))
USER_DATA = (os.path.join(DATA_FILES, "user_data.json"))
TASKS = (os.path.join(DATA_FILES, "tasks.json"))

WORKING_DAY_LENGTH = 480

FIRST_REPORT_ROW = 10

SOLID_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


# I'm terribly sorry for it but I can't find another way :( 
def dates_from_week_number(year, week_num):
    
    # Creating list of all days in the year 
    first_day = datetime.date(year, 1, 1)
    last_day = datetime.date(year, 12, 31)
    delta = datetime.timedelta(days=1)

    all_days_in_year = []
    
    current_day = first_day
    while current_day != last_day:
        all_days_in_year.append(current_day)
        current_day += delta
    all_days_in_year.append(last_day)
    
    # Creating list with dates of the needed week
    res = []
    for day in all_days_in_year:
        if day.isocalendar()[1] == week_num:
            res.append(day.strftime('%d/%m/%Y'))
        
    
    return res


def report_generator(week):
    
    task = {
        "task": "",
        "action":[],
        "time":0
    }
    
    res = []
    
    with open(TASKS, encoding="utf-8") as f:
        all_tasks = json.load(f)
        
    
    list_for_checking = []
    actions_list = []
    checking_action_list = []
    time = 0
    
    while len(res) < random.randint(5, 8): 
        random_num = random.randint(1, len(all_tasks)) 
        
        for i in range(1, len(all_tasks) + 1):        
            if week in NUM_WEEKS and all_tasks.get(f"task_{i}").get("weight") == 10:
                
                
                          
                random_num_actions = random.randint(1, len(all_tasks.get(f"task_{random_num}").get("actions")))
                while len(actions_list) < random_num_actions / 1.5:
                    for k in range(1, len(all_tasks.get(f"task_{random_num}").get("actions")) + 1):
                        
                        action_name = all_tasks.get(f"task_{random_num}").get("actions").get(f"action_{k}").get("name")
                        action_weight = all_tasks.get(f"task_{random_num}").get("actions").get(f"action_{k}").get("weight")
                        action_time = all_tasks.get(f"task_{random_num}").get("actions").get(f"action_{k}").get("time")
                        
                        
                        if action_weight == random.randint(1, action_weight):
                            if action_name not in checking_action_list:
                                actions_list.append(action_name)
                                time += action_time
                                pass
                            checking_action_list.append(action_name)
                                
                            pass
                        pass
                   
                task = {"task": all_tasks.get(f"task_{i}").get("name"),
                        "action": actions_list,
                        "time": time
                        }    
                    
                if task.get("task") not in list_for_checking:
                    res.append(task)
                    
                list_for_checking.append(task.get("task"))
       
           
        
         
        if all_tasks.get(f"task_{random_num}").get("weight") != 10:
            
            actions_list = []
            checking_action_list = []
            time = 0
                        
            random_num_actions = random.randint(1, len(all_tasks.get(f"task_{random_num}").get("actions")))
            while len(actions_list) < random_num_actions / 1.5:
                for k in range(1, len(all_tasks.get(f"task_{random_num}").get("actions")) + 1):
                    
                    action_name = all_tasks.get(f"task_{random_num}").get("actions").get(f"action_{k}").get("name")
                    action_weight = all_tasks.get(f"task_{random_num}").get("actions").get(f"action_{k}").get("weight")
                    action_time = all_tasks.get(f"task_{random_num}").get("actions").get(f"action_{k}").get("time")
                    
                    
                    if action_weight == random.randint(1, action_weight):
                        if action_name not in checking_action_list:
                            actions_list.append(action_name)
                            time += action_time
                            pass
                        checking_action_list.append(action_name)
                            
                        pass
                    pass
                
            task = {"task": all_tasks.get(f"task_{random_num}").get("name"),
                    "action": actions_list,
                    "time": time
                    }    
                
            if task.get("task") not in list_for_checking:
                res.append(task)
            list_for_checking.append(task.get("task"))        
        
    day_time = 0
      
    while day_time != WORKING_DAY_LENGTH:
        
        random_num_time = random.randint(0, len(res) - 1)
        if res[random_num_time].get("time") > 25:
            res[random_num_time].update({"time": res[random_num_time].get("time") - 5})
        day_time = 0
        for k in range(len(res)):
            day_time += res[k].get("time")
    
    
    
    return res
    
    
def report_writer():

    if int(combo_years.get()) not in YEARS:
        result.config(text="Не выбран год")
    
    if int(week_number.get()) not in range(1,53):
        result.config(text="Неправильный номер недели")
    
       
    path_to_reports = os.path.join(os.path.join(ROOT, "Отчёты"), combo_years.get())
    
    
    if not os.path.exists(path_to_reports):
        os.makedirs(path_to_reports)
        
    with open(USER_DATA, encoding="utf-8") as f:
        
        user_data = json.load(f)
        
    dates = dates_from_week_number(int(combo_years.get()), int(week_number.get()))
    
    start_data_list = dates[0].split("/")
    end_data_list = dates[-1].split("/")
    
    
    file_name = f"{start_data_list[0]}-{end_data_list[0]} {MONTHS[int(start_data_list[1]) - 1]} {end_data_list[2]}.xlsx"
    report_file = os.path.join(path_to_reports, file_name)  
    workbook = Workbook()
    worksheet = workbook.active
    
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 50
    worksheet.column_dimensions['C'].width = 60
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 15
    worksheet.row_dimensions[9].height = 45
        
    worksheet.merge_cells('A1:D1')
    worksheet["A1"] = "ОТЧЕТ"
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    worksheet.merge_cells('A2:D2')
    worksheet["A2"] = "о проделанной работе за неделю сотрудника на дистанционной работе"
    worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
    
    worksheet.merge_cells('A3:D3')
    worksheet["A3"] = f"Сотрудник: {user_data.get('Сотрудник')}"
    worksheet["A3"].alignment = Alignment(horizontal="left", vertical="center")
    
    worksheet.merge_cells('A4:D4')
    worksheet["A4"] = f"Наименование структурного подразделения: {user_data.get('Наименование структурного подразделения')}"
    worksheet["A4"].alignment = Alignment(horizontal="left", vertical="center")
    
    worksheet.merge_cells('A5:D5')
    worksheet["A5"] = f"Должность: {user_data.get('Должность')}"
    worksheet["A5"].alignment = Alignment(horizontal="left", vertical="center")
    
    dates = dates_from_week_number(int(combo_years.get()), int(week_number.get()))
    
    start_data_list = dates[0].split("/")
    end_data_list = dates[-1].split("/")
    
    worksheet.merge_cells('A6:D6')
    start_report = f"«{start_data_list[0]}» {MONTHS[int(start_data_list[1]) - 1]} {start_data_list[2]} г."
    end_report = f"«{end_data_list[0]}» {MONTHS[int(end_data_list[1]) - 1]} {end_data_list[2]} г."
    worksheet["A6"] = f"За период: с {start_report} по {end_report}"
    worksheet["A6"].alignment = Alignment(horizontal="left", vertical="center")
    
    worksheet.merge_cells('A7:D7')
    worksheet["A7"] = f"Место: {user_data.get('Место')}"
    worksheet["A7"].alignment = Alignment(horizontal="left", vertical="center")
    
    worksheet["A9"] = "Дата"
    worksheet["B9"] = "Задание"
    worksheet["C9"] = "Характеристика выполненной\nработы"
    worksheet["D9"] = "Длительность\nисполнения\nв день (мин)"
    worksheet["E9"] = "Время работы\n(ч. в день)"
    worksheet["A9"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    worksheet["B9"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    worksheet["C9"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    worksheet["D9"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    worksheet["E9"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    worksheet["A9"].border = SOLID_BORDER
    worksheet["B9"].border = SOLID_BORDER
    worksheet["C9"].border = SOLID_BORDER
    worksheet["D9"].border = SOLID_BORDER
    worksheet["E9"].border = SOLID_BORDER
    
    
    next_row = 9 # the last row of header
    
    for i in range(5):
        day_tasks = report_generator(week_number.get())
        
        worksheet.merge_cells(f'A{next_row + 1}:A{next_row + len(day_tasks)}')
        
        worksheet[f"A{next_row + 1}"] = dates[i]
        worksheet[f"A{next_row + 1}"].alignment = Alignment(horizontal="center", vertical="center")
        worksheet[f"A{next_row + 1}"].border = SOLID_BORDER
        
        for k in range(len(day_tasks)):
            
            worksheet[f"B{next_row + 1 + k}"] = day_tasks[k].get("task")
            worksheet[f"B{next_row + 1 + k}"].alignment = Alignment(horizontal="left", vertical="top")
            worksheet[f"B{next_row + 1 + k}"].border = SOLID_BORDER
            
            actions = ""
            
            for j in range(len(day_tasks[k].get("action"))):                
                actions += day_tasks[k].get("action")[j]
                actions += "\n"
                worksheet[f"C{next_row + 1 + k}"] = actions
                worksheet.row_dimensions[next_row + 1 + k].height = 15 * len(day_tasks[k].get("action"))
                worksheet[f"C{next_row + 1 + k}"].alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                worksheet[f"C{next_row + 1 + k}"].border = SOLID_BORDER
        
            day_time = 0
            day_time += day_tasks[k].get("time")
               
            worksheet[f"D{next_row + 1 + k}"] = day_time   
            worksheet[f"D{next_row + 1 + k}"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            worksheet[f"D{next_row + 1 + k}"].border = SOLID_BORDER
            
            
        worksheet.merge_cells(f'E{next_row + 1}:E{next_row + len(day_tasks)}')
        worksheet[f"E{next_row + 1}"] = 8
        worksheet[f"E{next_row + 1}"].alignment = Alignment(horizontal="center", vertical="center")
        #worksheet[f"E{next_row + 1}"].border = SOLID_BORDER
        #worksheet[f"E{next_row + len(day_tasks)}"].border = Border(bottom=Side(style="thin"))
        #worksheet[f"E{next_row + len(day_tasks)}"].border = SOLID_BORDER
        
        
        next_row += len(day_tasks)
        
        for next_cell in range(FIRST_REPORT_ROW, next_row + 1):
            
            worksheet[f"E{next_cell}"].border = SOLID_BORDER
            #worksheet[f"E{next_cell}"].border = Border(right=Side(style="thin"))
            pass
        
        pass
    worksheet[f"A{next_row + 1}"] = dates[-2]
    worksheet[f"A{next_row + 1}"].alignment = Alignment(horizontal="center", vertical="top")
    worksheet[f"A{next_row + 1}"].border = SOLID_BORDER
    worksheet[f"A{next_row + 2}"] = dates[-1]
    worksheet[f"A{next_row + 2}"].alignment = Alignment(horizontal="center", vertical="top")
    worksheet[f"A{next_row + 2}"].border = SOLID_BORDER
    
    worksheet[f"B{next_row + 1}"].border = SOLID_BORDER
    worksheet[f"D{next_row + 1}"].border = SOLID_BORDER
    worksheet[f"E{next_row + 1}"].border = SOLID_BORDER
    
    worksheet[f"C{next_row + 1}"] = "Выходной"
    worksheet[f"C{next_row + 1}"].alignment = Alignment(horizontal="center", vertical="top")
    worksheet[f"C{next_row + 1}"].border = SOLID_BORDER
    
    worksheet[f"C{next_row + 2}"] = "Выходной"
    worksheet[f"C{next_row + 2}"].alignment = Alignment(horizontal="center", vertical="top")
    worksheet[f"C{next_row + 2}"].border = SOLID_BORDER
    
    worksheet[f"B{next_row + 2}"].border = SOLID_BORDER
    worksheet[f"D{next_row + 2}"].border = SOLID_BORDER
    worksheet[f"E{next_row + 2}"].border = SOLID_BORDER
    
    
    
    
    worksheet.merge_cells(f'A{next_row + 5}:C{next_row + 5}')
    worksheet[f"A{next_row + 5}"] = user_data.get('Должность босса')
    worksheet[f"D{next_row + 5}"] = "______________"
    worksheet.merge_cells(f'E{next_row + 5}:F{next_row + 5}')
    worksheet[f"E{next_row + 5}"] = user_data.get('Босс')
    
    worksheet.merge_cells(f'A{next_row + 7}:C{next_row + 7}')
    worksheet[f"A{next_row + 7}"] = user_data.get('Должность')
    worksheet[f"D{next_row + 7}"] = "______________"
    worksheet.merge_cells(f'E{next_row + 7}:F{next_row + 7}')
    worksheet[f"E{next_row + 7}"] = user_data.get('Сотрудник')
    
    
    
    
        
    workbook.save(report_file)  
    
    
    result.config(text=f"Отчет за {week_number.get()}-ю неделю {combo_years.get()}-го года сформирован")
    pass



win = Tk()
win.resizable(False, False) 
win.title("Генератор отчетов")
win.geometry(f"{WIN_W}x{WIN_H}+{WIN_X}+{WIN_Y}")

Label(text="Выберите год", font=("Calibri", 16)).pack(pady=15)

combo_years = ttk.Combobox(win, values=YEARS)
combo_years.pack(pady=5)

Label(text="Введите номер недели (1-52)", font=("Calibri", 16)).pack(pady=15)

week_number = Entry()
week_number.pack(pady=5)


Button(text="Сформировать отчеты", command=report_writer).pack(pady=10)

result = Label()
result.pack(pady=20)

win.mainloop()

